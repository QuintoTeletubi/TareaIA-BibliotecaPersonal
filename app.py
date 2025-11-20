from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash
from models import db, User, Book
from datetime import datetime
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import os

# Configuración de la aplicación
app = Flask(__name__)

# Configuración para producción y desarrollo
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'tu_clave_secreta_super_segura_aqui')

# Base de datos: PostgreSQL para Render, SQLite para desarrollo
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Render PostgreSQL
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    # Desarrollo local
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///biblioteca.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Inicialización de extensiones
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder a esta página.'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Crear las tablas de la base de datos
with app.app_context():
    db.create_all()

@app.route('/')
def index():
    """Página principal - redirige al login si no está autenticado"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Registro de nuevos usuarios"""
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        
        # Verificar si el usuario ya existe
        if User.query.filter_by(username=username).first():
            flash('El nombre de usuario ya existe', 'error')
            return render_template('register.html')
        
        if User.query.filter_by(email=email).first():
            flash('El email ya está registrado', 'error')
            return render_template('register.html')
        
        # Crear nuevo usuario
        user = User(username=username, email=email)
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash('Registro exitoso. Por favor inicia sesión.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Inicio de sesión de usuarios"""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            flash('Inicio de sesión exitoso', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Cerrar sesión"""
    logout_user()
    flash('Has cerrado sesión exitosamente', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard principal con estadísticas y libros recientes"""
    books = Book.query.filter_by(user_id=current_user.id).order_by(Book.added_date.desc()).limit(5).all()
    
    # Estadísticas
    total_books = Book.query.filter_by(user_id=current_user.id).count()
    read_books = Book.query.filter_by(user_id=current_user.id, is_read=True).count()
    unread_books = total_books - read_books
    
    stats = {
        'total': total_books,
        'read': read_books,
        'unread': unread_books
    }
    
    return render_template('dashboard.html', books=books, stats=stats)

@app.route('/books')
@login_required
def books():
    """Lista completa de libros del usuario"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    status_filter = request.args.get('status', 'all')
    
    query = Book.query.filter_by(user_id=current_user.id)
    
    # Filtro por búsqueda
    if search:
        query = query.filter(
            (Book.title.contains(search)) |
            (Book.author.contains(search)) |
            (Book.genre.contains(search))
        )
    
    # Filtro por estado
    if status_filter == 'read':
        query = query.filter_by(is_read=True)
    elif status_filter == 'unread':
        query = query.filter_by(is_read=False)
    
    books = query.order_by(Book.added_date.desc()).paginate(
        page=page, per_page=10, error_out=False
    )
    
    return render_template('books.html', books=books, search=search, status_filter=status_filter)

@app.route('/book/add', methods=['GET', 'POST'])
@login_required
def add_book():
    """Agregar un nuevo libro"""
    if request.method == 'POST':
        title = request.form['title']
        author = request.form['author']
        genre = request.form['genre']
        is_read = 'is_read' in request.form
        rating = request.form.get('rating', type=int)
        comment = request.form.get('comment', '')
        read_date_str = request.form.get('read_date')
        
        # Convertir fecha si se proporciona
        read_date = None
        if read_date_str:
            try:
                read_date = datetime.strptime(read_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Formato de fecha inválido', 'error')
                return render_template('add_book.html')
        
        book = Book(
            title=title,
            author=author,
            genre=genre,
            is_read=is_read,
            rating=rating if rating else None,
            comment=comment,
            read_date=read_date,
            user_id=current_user.id
        )
        
        db.session.add(book)
        db.session.commit()
        
        flash('Libro agregado exitosamente', 'success')
        return redirect(url_for('books'))
    
    return render_template('add_book.html')

@app.route('/book/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_book(id):
    """Editar un libro existente"""
    book = Book.query.get_or_404(id)
    
    # Verificar que el libro pertenece al usuario actual
    if book.user_id != current_user.id:
        flash('No tienes permisos para editar este libro', 'error')
        return redirect(url_for('books'))
    
    if request.method == 'POST':
        book.title = request.form['title']
        book.author = request.form['author']
        book.genre = request.form['genre']
        book.is_read = 'is_read' in request.form
        book.rating = request.form.get('rating', type=int) or None
        book.comment = request.form.get('comment', '')
        
        read_date_str = request.form.get('read_date')
        if read_date_str:
            try:
                book.read_date = datetime.strptime(read_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Formato de fecha inválido', 'error')
                return render_template('edit_book.html', book=book)
        else:
            book.read_date = None
        
        db.session.commit()
        flash('Libro actualizado exitosamente', 'success')
        return redirect(url_for('books'))
    
    return render_template('edit_book.html', book=book)

@app.route('/book/<int:id>/delete', methods=['POST'])
@login_required
def delete_book(id):
    """Eliminar un libro"""
    book = Book.query.get_or_404(id)
    
    # Verificar que el libro pertenece al usuario actual
    if book.user_id != current_user.id:
        flash('No tienes permisos para eliminar este libro', 'error')
        return redirect(url_for('books'))
    
    db.session.delete(book)
    db.session.commit()
    flash('Libro eliminado exitosamente', 'success')
    
    return redirect(url_for('books'))

@app.route('/export/excel')
@login_required
def export_excel():
    """Exportar biblioteca a Excel"""
    books = Book.query.filter_by(user_id=current_user.id).all()
    
    if not books:
        flash('No tienes libros para exportar', 'warning')
        return redirect(url_for('books'))
    
    # Convertir libros a lista de diccionarios
    data = [book.to_dict() for book in books]
    
    # Crear archivo Excel con openpyxl
    filename = f'biblioteca_{current_user.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join('exports', filename)
    
    # Asegurar que el directorio existe
    os.makedirs('exports', exist_ok=True)
    
    # Crear workbook y worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mi Biblioteca'
    
    # Escribir encabezados
    if data:
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Escribir datos
        for row, book_data in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=book_data[header])
    
    # Guardar archivo
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route('/export/pdf')
@login_required
def export_pdf():
    """Exportar biblioteca a PDF"""
    books = Book.query.filter_by(user_id=current_user.id).all()
    
    if not books:
        flash('No tienes libros para exportar', 'warning')
        return redirect(url_for('books'))
    
    # Crear archivo PDF
    filename = f'biblioteca_{current_user.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    filepath = os.path.join('exports', filename)
    
    # Asegurar que el directorio existe
    os.makedirs('exports', exist_ok=True)
    
    # Crear documento PDF
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    # Título
    title = Paragraph(f"Biblioteca Personal de {current_user.username}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Preparar datos para la tabla
    data = [['Título', 'Autor', 'Género', 'Estado', 'Calificación']]
    
    for book in books:
        status = 'Leído' if book.is_read else 'Por leer'
        rating = f'{book.rating}/5' if book.rating else 'N/A'
        data.append([
            book.title[:30] + '...' if len(book.title) > 30 else book.title,
            book.author[:20] + '...' if len(book.author) > 20 else book.author,
            book.genre,
            status,
            rating
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[2*inch, 1.5*inch, 1*inch, 0.8*inch, 0.7*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

# Inicializar base de datos automáticamente
with app.app_context():
    try:
        db.create_all()
        # Crear usuario demo si no existe
        if not User.query.filter_by(username='demo').first():
            demo_user = User(
                username='demo',
                email='demo@biblioteca.com',
                password_hash=generate_password_hash('demo123')
            )
            db.session.add(demo_user)
            db.session.commit()
            print("Usuario demo creado: demo/demo123")
    except Exception as e:
        print(f"Error inicializando BD: {e}")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)